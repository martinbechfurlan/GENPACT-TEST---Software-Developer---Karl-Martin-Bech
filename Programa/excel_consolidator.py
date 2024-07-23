import os
import shutil
import time
from openpyxl import Workbook, load_workbook

class ExcelConsolidator:
    def __init__(self, master_file, log_callback):
        # Initialize the class with the path to the master file and a logging callback
        self.master_file = master_file
        self.log_callback = log_callback
        # If the master file does not exist, create it
        if not os.path.exists(self.master_file):
            wb = Workbook()  # Create a new Workbook
            wb.save(self.master_file)  # Save it to the specified path
            self.log_callback(f"Created master file: {self.master_file}")

    def consolidate(self, file_path):
        # Extract the base name of the file (without extension)
        file_name = os.path.splitext(os.path.basename(file_path))[0]

        # Create the 'processed' folder if it doesn't exist
        processed_folder = os.path.join(os.path.dirname(file_path), 'processed')
        os.makedirs(processed_folder, exist_ok=True)
        # Check if the file has already been processed
        if self.is_file_processed(file_path, processed_folder):
            # Move the file to 'processed' with a copy suffix if it has been processed
            new_file_path = self.get_unique_file_path(processed_folder, file_name, os.path.splitext(file_path)[1])
            shutil.move(file_path, new_file_path)
            self.log_callback(f"File already processed. Moved {file_path} to processed folder as {new_file_path}")
            return

        # Add sheets from the new file to the master file
        while True:
            if not self._is_temporary_file(self.master_file):
                try:
                    wb_master = load_workbook(self.master_file)  # Load the master workbook
                    break
                except PermissionError:
                    # If the master file is locked, wait and retry
                    self.log_callback(f"Waiting for {self.master_file} to be released...")
                    time.sleep(1)  # Wait for 1 second before retrying
            else:
                time.sleep(1)  # Wait for 1 second before retrying

        # Load the new workbook
        wb_new = load_workbook(file_path)

        for sheet_name in wb_new.sheetnames:
            sheet = wb_new[sheet_name]
            new_sheet_name = f"{file_name}_{sheet_name}"
            # Ensure unique sheet name in the master workbook
            if new_sheet_name in wb_master.sheetnames:
                new_sheet_name = self.generate_unique_sheet_name(wb_master, new_sheet_name)
            new_sheet = wb_master.create_sheet(new_sheet_name)
            # Copy rows from the new sheet to the master sheet
            for row in sheet.iter_rows(values_only=True):
                new_sheet.append(row)

        while True:
            if not self._is_temporary_file(self.master_file):
                try:
                    wb_master.save(self.master_file)  # Save changes to the master workbook
                    break
                except PermissionError:
                    # If the master file is locked, wait and retry
                    self.log_callback(f"Waiting for {self.master_file} to be released...")
                    time.sleep(1)  # Wait for 1 second before retrying
            else:
                time.sleep(1)  # Wait for 1 second before retrying

        self.log_callback(f"Added {file_path} to master file")
        # Move the processed file to the 'processed' folder
        self.move_to_processed(file_path)

    def _is_temporary_file(self, file_path):
        # Check if the file is a temporary file (starts with '~$')
        return os.path.basename(file_path).startswith('~$')

    def generate_unique_sheet_name(self, workbook, base_name):
        # Generate a unique sheet name by appending a counter if necessary
        counter = 1
        new_name = f"{base_name}_{counter}"
        while new_name in workbook.sheetnames:
            counter += 1
            new_name = f"{base_name}_{counter}"
        return new_name

    def move_to_processed(self, file_path):
        # Move the file to the 'processed' folder
        processed_folder = os.path.join(os.path.dirname(file_path), 'processed')
        os.makedirs(processed_folder, exist_ok=True)
        self._move_file(file_path, processed_folder, "processed")

    def move_to_not_applicable(self, file_path):
        # Move the file to the 'not_applicable' folder
        not_applicable_folder = os.path.join(os.path.dirname(file_path), 'not_applicable')
        os.makedirs(not_applicable_folder, exist_ok=True)
        # Check if the file already exists and rename it if necessary
        new_file_path = self.get_unique_file_path(not_applicable_folder, os.path.splitext(os.path.basename(file_path))[0], os.path.splitext(file_path)[1])
        shutil.move(file_path, new_file_path)
        self.log_callback(f"Moved {file_path} to not applicable folder as {new_file_path}")

    def _move_file(self, file_path, target_folder, folder_name):
        # Move the file to the target folder with retry logic
        max_retries = 10
        time.sleep(1)  # Wait for 1 second before trying
        for attempt in range(max_retries):
            try:
                shutil.move(file_path, target_folder)
                self.log_callback(f"Moved {file_path} to {folder_name} folder")
                break
            except PermissionError as e:
                # Handle permission errors by retrying
                self.log_callback(f"Error moving {file_path} to {folder_name} folder: {e}")
                self.log_callback(f"Retrying ({attempt + 1}/{max_retries})...")
                time.sleep(1)  # Wait for 1 second before retrying
            except Exception as e:
                # Handle any unexpected errors
                self.log_callback(f"Unexpected error: {e}")
                break

    def get_unique_file_path(self, folder, base_name, ext):
        # Generate a unique file path by appending a counter if necessary
        counter = 1
        new_file_name = f"{base_name}{ext}"
        new_file_path = os.path.join(folder, new_file_name)

        while os.path.exists(new_file_path):
            new_file_name = f"{base_name}_copy{counter}{ext}"
            new_file_path = os.path.join(folder, new_file_name)
            counter += 1

        return new_file_path

    def is_file_processed(self, file_path, processed_folder):
        # Check if the file has already been processed
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        ext = os.path.splitext(file_path)[1]
        return any(f.startswith(base_name) and f.endswith(ext) for f in os.listdir(processed_folder))
